import pandas as pd
import re
from bs4 import BeautifulSoup
import ast
import html
from openpyxl import load_workbook
from openpyxl.styles import Font

# Load files
metadata_df = pd.read_csv('../data/unified_mannheim_metadata.csv')

# Precompiled regex patterns
_RE_CTRL          = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
_RE_BSL_ESC       = re.compile(r"\\[nrt]")
_RE_ZWSP          = re.compile(r"[\u200B-\u200F\u202A-\u202E\u2060]")
_RE_MD_HEADINGS   = re.compile(r"(?m)^\s*#{1,6}\s*")
_RE_MD_EMPH       = re.compile(r"[*_~`]+")
_RE_WS            = re.compile(r"\s+")
_RE_LISTY_STR     = re.compile(r"^\s*\[\s*['\"]")
_RE_DOI           = re.compile(r"(10\.\d{4,9}/\S+)", re.IGNORECASE)
_MAX_LEN_DEFAULT  = 10000

_MISSING_TOKEN = "NA"

# Help functions
def _normalize_input(text):
    if text is None or pd.isna(text):
        return None

    if isinstance(text, (bytes, bytearray)):
        try:
            text = text.decode("utf-8", "replace")
        except Exception:
            text = str(text)

    return str(text)

def _apply_missing_token_if_needed(s: str) -> str:
    if s is None:
        return _MISSING_TOKEN

    if _is_missing_after_clean(s):
        return _MISSING_TOKEN

    if s.strip().lower() == "nan":
        return _MISSING_TOKEN

    return s

def _unwrap_listlike_string(s: str) -> str:
    if _RE_LISTY_STR.match(s) and s.rstrip().endswith("]"):
        try:
            parts = ast.literal_eval(s)
            if isinstance(parts, (list, tuple)) and parts:
                first = str(parts[0]).strip()
                if len(first) <= 10 and len(parts) > 1:
                    return str(parts[1]).strip()
                return first
        except Exception:
            return s
    return s

def _decode_html_and_strip_tags(s: str) -> str:
    s = html.unescape(s)
    s = BeautifulSoup(s, "html.parser").get_text(separator=" ")
    return s

def _remove_backslash_escapes(s: str) -> str:
    return _RE_BSL_ESC.sub(" ", s)

def _remove_control_and_zwsp(s: str) -> str:
    s = _RE_CTRL.sub(" ", s)
    s = _RE_ZWSP.sub("", s)
    return s

def _strip_markdown_light(s: str) -> str:
    s = _RE_MD_HEADINGS.sub("", s)
    s = _RE_MD_EMPH.sub("", s)
    return s

def _collapse_whitespace(s: str) -> str:
    return _RE_WS.sub(" ", s).strip()

# Cleaning functions
def _is_missing_after_clean(s: str) -> bool:
    t = s.strip()
    return t == "" or t == "-"

def clean_description(text, *, max_len: int = _MAX_LEN_DEFAULT) -> str:
    s = _normalize_input(text)
    if s is None:
        return _MISSING_TOKEN

    s = s.strip()
    s = _apply_missing_token_if_needed(s)
    if s == _MISSING_TOKEN:
        return s

    s = _unwrap_listlike_string(s)

    s = _decode_html_and_strip_tags(s)

    s = _remove_backslash_escapes(s)
    s = _remove_control_and_zwsp(s)

    s = _strip_markdown_light(s)

    s = _collapse_whitespace(s)

    s = _apply_missing_token_if_needed(s)
    if s == _MISSING_TOKEN:
        return s

    if max_len and len(s) > max_len:
        s = s[:max_len].rstrip() + "…"

    return s

def clean_title(text) -> str:
    s = _normalize_input(text)
    if s is None:
        return _MISSING_TOKEN

    s = s.replace("\n", " ").replace("\r", " ").replace("\t", " ")

    s = _remove_control_and_zwsp(s)

    s = _collapse_whitespace(s)

    s = _apply_missing_token_if_needed(s)

    return s

def clean_doi(value) -> str:
    if value is None or pd.isna(value):
        return _MISSING_TOKEN

    s = str(value).strip()
    if not s:
        return _MISSING_TOKEN

    match = _RE_DOI.search(s)
    if not match:
        return _MISSING_TOKEN

    doi = match.group(1)

    # Remove unwanted trailing characters
    doi = doi.rstrip("',;:\" ")

    if not doi:
        return _MISSING_TOKEN

    return f"https://doi.org/{doi}"

# Apply cleaning functions
metadata_df["Description"] = metadata_df["Description"].apply(clean_description)
metadata_df["Title"] = metadata_df["Title"].apply(clean_title)
metadata_df["DOI"] = metadata_df["DOI"].apply(clean_doi)

# Reorder cols
desired_order = [
    "DOI",
    "Type",
    "Title",
    "Creators",
    "Affiliations",
    "Description",
    "Year",
    "License"
]

existing_order = [col for col in desired_order if col in metadata_df.columns]
remaining_cols = [col for col in metadata_df.columns if col not in existing_order]
metadata_df = metadata_df[existing_order + remaining_cols]

# Ensure NA is consistent
metadata_df["Description"] = metadata_df["Description"].fillna(_MISSING_TOKEN)

# Save to CSV
metadata_df.to_csv('../data/unified_mannheim_metadata_cleaned.csv', index=False)
# Save to JSON
metadata_df.to_json("../data/unified_mannheim_metadata_cleaned.json", orient="records")

# Save to XLSX with hyperlink objects for col DOI
excel_path = '../data/unified_mannheim_metadata_cleaned.xlsx'
metadata_df.to_excel(excel_path, index=False, sheet_name='Datenbibliografie')

wb = load_workbook(excel_path)
ws = wb['Datenbibliografie']

doi_col_idx = metadata_df.columns.get_loc("DOI") + 1
link_font = Font(color="0000EE", underline="single")

for row_idx in range(2, ws.max_row + 1):  # row 1 = header
    cell = ws.cell(row=row_idx, column=doi_col_idx)
    value = cell.value

    if isinstance(value, str) and value and value != _MISSING_TOKEN:
        cell.hyperlink = value
        cell.font = link_font

wb.save(excel_path)


print("Cleaned unified metadata saved to ../data/unified_mannheim_metadata_cleaned.csv")
